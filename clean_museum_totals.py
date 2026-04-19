#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Construit la version finale nettoyee du fichier musee.

Le resultat final doit contenir exactement ces colonnes :
- Items
- Stack max
- Quantité requise
- english_name
- Total Pour Item Craftables
- Total à Farm
- Equialent Coffre
"""

from __future__ import annotations

import math
from pathlib import Path

from openpyxl import Workbook, load_workbook


SOURCE_FILE = Path("Musee_Infinis_clean_with_totals.xlsx")
OUTPUT_DIR = Path("sortie final clean")
OUTPUT_FILE = OUTPUT_DIR / "Musee_Infinis_clean_with_totals_clean.xlsx"


def normalize_bool_text(value: object) -> str:
    """
    Convertit une valeur Excel en texte `true` / `false`.

    Cela permet de comparer proprement des cellules qui peuvent contenir :
    - un booléen Python
    - une chaîne de caractères
    - une cellule vide
    """
    if value is True:
        return "true"
    if value is False:
        return "false"
    return str(value).strip().lower()


def to_int(value: object) -> int:
    """
    Convertit une valeur Excel en entier.

    On sécurise la conversion parce que les cellules Excel peuvent contenir
    `None`, du texte ou des nombres déjà convertis en float.
    """
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def clean_museum_file(source_path: Path, output_path: Path) -> None:
    """
    Lit le fichier source enrichi et reconstruit le tableau final.

    Logique appliquée :
    1. lire les colonnes utiles du fichier source
    2. si `OnRecipes_Res` et `OnRecipes_Ingredient` sont tous les deux `false`,
       alors `Total à Farm` doit être remis à 0
    3. construire `Total à Farm` à partir de :
       `Total Pour Item Craftables` + `Total cible item`
    4. calculer `Equialent Coffre` avec la règle demandée :
       - 167.0 -> 167
       - 167.00001 -> 168
       - 167.796875 -> 168
    5. écrire un nouveau fichier Excel propre, sans modifier le fichier source
    """
    if not source_path.exists():
        raise FileNotFoundError(f"Fichier introuvable: {source_path}")

    # On crée le dossier final automatiquement pour que le script puisse
    # être relancé sans préparation manuelle.
    output_path.parent.mkdir(parents=True, exist_ok=True)

    source_workbook = load_workbook(source_path, data_only=True)
    try:
        sheet_name = "Sheet1" if "Sheet1" in source_workbook.sheetnames else source_workbook.sheetnames[0]
        source_worksheet = source_workbook[sheet_name]

        # On crée un nouveau classeur vide au lieu de réutiliser celui chargé.
        # C'est plus fiable pour produire un fichier final propre et évite
        # les effets de bord quand on supprime/reconstruit des colonnes.
        output_workbook = Workbook()
        output_worksheet = output_workbook.active
        output_worksheet.title = sheet_name

        headers = [
            "Items",
            "Stack max",
            "Quantité requise",
            "english_name",
            "Total Pour Item Craftables",
            "Total à Farm",
            "Equialent Coffre",
        ]
        output_worksheet.append(headers)

        cleaned_rows = 0

        # Chaque ligne du fichier source est relue puis réécrite dans le format
        # final voulu. On garde donc la logique métier, mais on simplifie
        # complètement la structure du fichier de sortie.
        for row_number in range(2, source_worksheet.max_row + 1):
            item_name = source_worksheet.cell(row=row_number, column=1).value
            stack_max = source_worksheet.cell(row=row_number, column=2).value
            quantity_required = source_worksheet.cell(row=row_number, column=3).value
            english_name = source_worksheet.cell(row=row_number, column=7).value

            # Le fichier source contient déjà le total "pour items craftables".
            total_for_item_craftables = to_int(source_worksheet.cell(row=row_number, column=8).value)

            # Le "Total à Farm" attendu dans le fichier final correspond au
            # total craftable + la quantité cible de l'item lui-même.
            total_target_item = to_int(source_worksheet.cell(row=row_number, column=9).value)
            total_to_farm = total_for_item_craftables + total_target_item

            on_recipes_res = normalize_bool_text(source_worksheet.cell(row=row_number, column=10).value)
            on_recipes_ingredient = normalize_bool_text(source_worksheet.cell(row=row_number, column=11).value)

            # Si l'item n'apparait dans aucune recette, il ne doit pas compter
            # comme un besoin de farm pour les items craftables.
            if on_recipes_res == "false" and on_recipes_ingredient == "false":
                total_to_farm = 0
                cleaned_rows += 1

            quantity_required_value = to_int(quantity_required)
            if quantity_required_value <= 0:
                equivalent_chest = 0
            else:
                # On divise bien "Total à Farm" par "Quantité requise".
                # `math.ceil` applique exactement la règle voulue :
                # 167.0 devient 167, mais 167.00001 devient 168.
                equivalent_chest = math.ceil(total_to_farm / quantity_required_value)

            output_worksheet.append(
                [
                    item_name,
                    stack_max,
                    quantity_required,
                    english_name,
                    total_for_item_craftables,
                    total_to_farm,
                    equivalent_chest,
                ]
            )

        output_workbook.save(output_path)
        output_workbook.close()
    finally:
        source_workbook.close()

    print(f"[Termine] Fichier cree : {output_path}")
    print(f"[Termine] Lignes nettoyees : {cleaned_rows}")


if __name__ == "__main__":
    clean_museum_file(SOURCE_FILE, OUTPUT_FILE)
