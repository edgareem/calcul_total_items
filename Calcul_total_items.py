#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Analyse des besoins de farm pour remplir un coffre de 27 slots pour chaque item.

Entrées :
- items.json
- recipes.json

Hypothèses :
- Les deux fichiers sont à la racine, au même niveau que ce script.
- On veut remplir 27 slots pour chaque item.
- La quantité cible d'un item = 27 * stackSize.
- On utilise une liste manuelle de ressources "base farmables".
- On choisit la recette qui minimise le coût en ressources de base.
- Les recettes réversibles sont gérées naturellement, avec protection anti-cycle.

Sorties :
- farming_summary.json : résultat complet
- farming_totals.txt    : résumé lisible
"""

from __future__ import annotations

# ---------------------------------------------------------------------------
# GUIDE DE LECTURE POUR DEBUTANT
# ---------------------------------------------------------------------------
# Ce script est volontairement decompose en petites fonctions.
# L'idee est de separer chaque responsabilite :
#
# 1. Charger les donnees
#    - lire les JSON
#    - lire le fichier Excel
#
# 2. Transformer les donnees
#    - creer des dictionnaires de recherche rapide
#    - convertir les recettes dans un format plus simple a manipuler
#
# 3. Resoudre un besoin pour un item
#    - "de quoi ai-je besoin pour fabriquer cet item ?"
#    - si un ingredient a lui-meme une recette, on recommence
#    - si on arrive a une ressource de base, on s'arrete
#
# 4. Agreger les resultats
#    - total global de ce qu'il faut farmer
#    - detail par item
#    - fichiers de sortie lisibles
#
# Pourquoi autant de dictionnaires ?
# - un dictionnaire permet de retrouver une valeur tres vite a partir d'une cle
# - ici on cherche tres souvent un item par son id ou par son nom
#
# Pourquoi Counter ?
# - Counter est pratique pour additionner des quantites d'objets
# - exemple : 3 stick + 5 stick = 8 stick automatiquement
#
# Pourquoi de la recursion ?
# - une recette peut demander des ingredients
# - ces ingredients peuvent eux-memes demander d'autres ingredients
# - la structure est naturellement "en arbre"
#
# Pourquoi memoiser ?
# - si on a deja calcule le cout d'un item pour une certaine quantite,
#   inutile de refaire le travail
# - cela accelere enormement le programme
# ---------------------------------------------------------------------------

import json
import math
from collections import Counter, defaultdict
from copy import deepcopy
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Any

from openpyxl import load_workbook


# =========================
# Configuration utilisateur
# =========================

CHEST_SLOTS = 27

# Liste des ressources qu'on considère comme "à farmer".
# Dès qu'on atteint un de ces items, on s'arrête.
BASE_FARMABLES_BY_NAME = {
    # Bois
    "oak_log", "spruce_log", "birch_log", "jungle_log", "acacia_log",
    "dark_oak_log", "mangrove_log", "cherry_log", "pale_oak_log",
    "crimson_stem", "warped_stem",

    # Pierre / terre / sable
    "cobblestone", "blackstone", "basalt", "sand", "red_sand",
    "gravel", "clay_ball", "netherrack", "end_stone", "obsidian",

    # Minerais / ressources brutes
    "raw_iron", "raw_copper", "raw_gold", "coal", "charcoal",
    "redstone", "lapis_lazuli", "diamond", "emerald", "quartz",
    "amethyst_shard", "flint",

    # Mob / loot / organique
    "bone", "string", "slime_ball", "gunpowder", "leather",
    "feather", "egg", "wheat", "sugar_cane", "melon_slice",
    "pumpkin", "cocoa_beans", "honeycomb", "bamboo", "stick",

    # Autres utiles
    "netherite_scrap", "ancient_debris", "blaze_rod", "ender_pearl",
}

# Regles de normalisation des ressources :
# - on ramene plusieurs formes d'une meme ressource vers une forme canonique
# - on evite de planifier des conversions reversibles inutiles
# - le ratio est applique ainsi :
#   canonical_qty = ceil(alias_qty * multiplier_num / multiplier_den)
# Exemples :
# - 1 iron_block -> 9 iron_ingot
# - 9 iron_nugget -> 1 iron_ingot
# - 8 stick -> 1 oak_log
# Pour les sticks, on force la chaine :
# stick -> oak_planks -> oak_log
# afin d'exprimer la recolte en buches de chene.

PRIORITY_BASE_ITEMS_BY_NAME = {
    "iron_ingot", "gold_ingot", "copper_ingot", "netherite_ingot",
    "diamond", "emerald", "redstone", "lapis_lazuli", "coal", "quartz", 
    "oak_log", "spruce_log", "birch_log", "jungle_log",
    "acacia_log", "dark_oak_log", "mangrove_log", "cherry_log",
    "crimson_stem", "warped_stem",
}

# alias_name -> (canonical_name, multiplier_num, multiplier_den)
REVERSIBLE_CANONICAL_RULES = {
    "iron_nugget": ("iron_ingot", 1, 9),
    "iron_block": ("iron_ingot", 9, 1),
    "gold_nugget": ("gold_ingot", 1, 9),
    "gold_block": ("gold_ingot", 9, 1),
    "copper_block": ("copper_ingot", 9, 1),
    "netherite_block": ("netherite_ingot", 9, 1),
    "diamond_block": ("diamond", 9, 1),
    "emerald_block": ("emerald", 9, 1),
    "redstone_block": ("redstone", 9, 1),
    "lapis_block": ("lapis_lazuli", 9, 1),
    "coal_block": ("coal", 9, 1),
    "quartz_block": ("quartz", 4, 1),
    "stick": ("oak_log", 1, 8),
}

# Ajustements manuels ajoutes au total global des ressources a farmer.
# Format : item_name -> qty a additionner au total final.
MANUAL_TOTAL_ADJUSTMENTS_BY_NAME = {
    #"arrow": 77760,
    "flint": 19440,
    "feather": 19440,
    #"glass_bottle": 3672,
    "sand": 3672,

    "oak_log": 4968,
    "leather": 1161,
    "sugar_cane": 3483,
}

# Items à exclure complètement du calcul, même s'ils ne viennent pas du fichier Excel
STATIC_EXCLUDED_ITEMS_BY_NAME = {
    "farmland",   # pas un item normal de stockage
}

ITEMS_FILE = Path("items.json")
RECIPES_FILE = Path("recipes.json")
SELECTED_RECIPES_FILE = Path("selected_recipe_choices.json")
MUSEUM_EXCEL_FILE = Path("Musee_Infinis_clean.xlsx")
MUSEUM_OUTPUT_EXCEL_FILE = Path("Musee_Infinis_clean_with_totals.xlsx")
OUTPUT_JSON = Path("farming_summary.json")
OUTPUT_TXT = Path("farming_totals.txt")
OUTPUT_TREE_TXT = Path("farming_recipe_trees.txt")

VERBOSE_EXCEL_EVERY = 1000


# =========================
# Chargement des données
# =========================

def load_json(path: Path) -> Any:
    """
    Lit un fichier JSON et retourne son contenu Python.

    Pourquoi Path.open(...) ?
    - `Path` vient de `pathlib` et rend la gestion des chemins plus lisible.
    - `encoding="utf-8"` force un encodage explicite, ce qui evite des
      problemes selon la machine ou le terminal.

    Pourquoi `json.load(...)` ?
    - parce qu'on lit directement depuis un fichier ouvert
    - si on avait deja le texte en memoire, on utiliserait `json.loads(...)`
    """
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def build_item_maps(items_data: List[Dict[str, Any]]) -> Tuple[Dict[int, dict], Dict[str, dict]]:
    """
    Construit deux dictionnaires de recherche rapide a partir de la liste des items.

    Sortie :
    - by_id[item_id] = item complet
    - by_name[item_name] = item complet

    Pourquoi faire cela ?
    - `items.json` est une liste
    - chercher un item dans une liste oblige a parcourir tous les elements
    - avec un dictionnaire, l'acces par cle est direct et bien plus rapide
    """
    by_id = {}
    by_name = {}

    # On parcourt toute la liste une seule fois pour preparer deux index.
    for item in items_data:
        by_id[item["id"]] = item
        by_name[item["name"]] = item
    return by_id, by_name


def normalize_item_lookup_key(value: str) -> str:
    """
    Normalise un texte pour faciliter les correspondances.

    Exemple :
    - "Oak Log" devient "oak_log"
    - " oak log " devient aussi "oak_log"

    Cette fonction est utile quand les sources de donnees n'ecrivent pas
    exactement les noms de la meme facon.
    """
    return str(value).strip().lower().replace(" ", "_")


def build_item_lookup_map(items_data: List[Dict[str, Any]]) -> Dict[str, str]:
    """
    Construit une table de correspondance tres permissive vers `item["name"]`.

    On accepte plusieurs formes d'un meme item :
    - son `name` technique
    - son `displayName`
    - une version normalisee de chacun

    Exemple :
    - "Oak Log"
    - "oak_log"
    - "oak log"
    pointent tous vers `oak_log`.
    """
    lookup = {}
    for item in items_data:
        item_name = item.get("name")
        display_name = item.get("displayName")
        if not item_name:
            continue

        # On enregistre plusieurs variantes de la meme cle pour etre tolerant
        # avec les differences d'ecriture entre les fichiers.
        for candidate in (item_name, display_name, normalize_item_lookup_key(item_name)):
            if candidate:
                lookup.setdefault(str(candidate).strip(), item_name)
                lookup.setdefault(normalize_item_lookup_key(str(candidate)), item_name)

        if display_name:
            display_as_name = normalize_item_lookup_key(display_name)
            lookup.setdefault(display_as_name, item_name)

    return lookup


def load_allowed_items_from_excel(
    excel_path: Path,
    items_by_name: Dict[str, dict],
    item_lookup_map: Dict[str, str],
) -> Set[str]:
    """
    Lit le fichier Excel du musee et retourne l'ensemble des items autorises.

    On utilise ici `openpyxl`, une bibliotheque tres pratique pour lire et
    ecrire des fichiers Excel `.xlsx` depuis Python.

    Choix importants :
    - `read_only=True` :
      on lit le fichier en mode lecture seule pour consommer moins de memoire
    - `data_only=True` :
      si certaines cellules contiennent des formules, on lit leur valeur
      calculee plutot que la formule brute

    La fonction retourne un `set` plutot qu'une liste car :
    - on veut des elements uniques
    - les tests `x in mon_set` sont tres rapides
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Fichier introuvable: {excel_path}")

    print(f"[Excel] Ouverture du fichier : {excel_path}", flush=True)
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet_name = "Sheet1" if "Sheet1" in workbook.sheetnames else workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        print(f"[Excel] Feuille utilisee : {sheet_name}", flush=True)

        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if header_row is None:
            raise ValueError("Impossible de lire la ligne d'entetes du fichier Excel.")

        headers = {
            str(value).strip().lower(): index
            for index, value in enumerate(header_row)
            if value is not None
        }

        english_name_col = headers.get("english_name")

        if english_name_col is None:
            raise ValueError("Colonne Excel introuvable: 'english_name'.")

        allowed_items = set()
        print(f"[Excel] Debut de lecture des lignes de donnees a conserver...", flush=True)

        # `iter_rows(..., values_only=True)` retourne directement les valeurs
        # des cellules, sans creer des objets Cell complets. C'est plus rapide
        # et largement suffisant pour notre besoin.
        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            excel_item_name = row_values[english_name_col] if english_name_col < len(row_values) else None

            if row_number == 5 or row_number % VERBOSE_EXCEL_EVERY == 0:
                print(
                    f"[Excel] Ligne {row_number}: english_name={excel_item_name!r}",
                    flush=True,
                )

            # Si la cellule est vide, il n'y a rien a exploiter sur cette ligne.
            if not excel_item_name:
                continue

            raw_excel_item_name = str(excel_item_name).strip()
            normalized_excel_item_name = normalize_item_lookup_key(raw_excel_item_name)
            resolved_name = (
                item_lookup_map.get(raw_excel_item_name)
                or item_lookup_map.get(normalized_excel_item_name)
            )
            if resolved_name and resolved_name in items_by_name:
                allowed_items.add(resolved_name)
                print(
                    f"[Excel] Item conserve ligne {row_number}: {excel_item_name} -> {resolved_name}",
                    flush=True,
                )
            else:
                print(
                    f"[Excel] Item non resolu ligne {row_number}: {excel_item_name!r}",
                    flush=True,
                )

        # On retire enfin quelques exclusions "hardcodees" qui ne doivent
        # jamais entrer dans le calcul, meme si elles apparaissent ailleurs.
        allowed_items.difference_update(STATIC_EXCLUDED_ITEMS_BY_NAME)
        print(f"[Excel] Lecture terminee. Total conserves: {len(allowed_items)}", flush=True)
        return allowed_items
    finally:
        workbook.close()


def print_progress(current: int, total: int, width: int = 40) -> None:
    """
    Affiche une barre de progression dans le terminal.

    La technique utilise `\\r` (retour chariot) pour reecrire la meme ligne
    plutot que d'en afficher une nouvelle a chaque iteration.
    """
    if total <= 0:
        return

    progress_ratio = current / total
    filled = int(width * progress_ratio)
    bar = "#" * filled + "-" * (width - filled)
    print(f"\rProgression: [{bar}] {current}/{total} ({progress_ratio:.1%})", end="", flush=True)

    if current >= total:
        print()


# =========================
# Parsing des recettes
# =========================

def extract_ingredients_from_recipe(recipe: Dict[str, Any]) -> Counter:
    """
    Extrait les ingredients d'une recette et les compte avec un Counter.

    Un Counter est une structure tres pratique pour dire :
    - cet ingredient apparait 1 fois
    - cet autre ingredient apparait 3 fois

    Cela evite de gerer manuellement des additions dans un dictionnaire.
    """
    """
    Retourne un Counter {item_id: quantité} pour une recette.

    Gère :
    - inShape : grille avec répétitions des ids
    - ingredients : liste d'ingrédients
    """
    ingredients = Counter()

    if "inShape" in recipe and recipe["inShape"] is not None:
        for row in recipe["inShape"]:
            for cell in row:
                if cell is not None:
                    ingredients[cell] += 1

    elif "ingredients" in recipe and recipe["ingredients"] is not None:
        for ing in recipe["ingredients"]:
            if ing is not None:
                ingredients[ing] += 1

    return ingredients


def normalize_recipes(raw_recipes: Dict[str, List[Dict[str, Any]]]) -> Dict[int, List[Dict[str, Any]]]:
    """
    Reorganise les recettes dans une structure plus simple pour le reste du code.

    Le JSON brut est correct, mais pas ideal pour faire beaucoup de recherches.
    Ici on prepare une version "pre-analysee" des recettes.
    """
    """
    Transforme recipes.json en structure plus simple :
    {
      result_item_id: [
        {
          "result_id": int,
          "result_count": int,
          "ingredients": Counter({ingredient_id: qty, ...}),
          "raw": recipe_originale
        },
        ...
      ]
    }
    """
    normalized = defaultdict(list)

    for result_id_str, recipe_list in raw_recipes.items():
        try:
            result_id_from_key = int(result_id_str)
        except ValueError:
            continue

        if not isinstance(recipe_list, list):
            continue

        for recipe in recipe_list:
            result = recipe.get("result")
            if not result or "id" not in result:
                continue

            result_id = result["id"]
            result_count = int(result.get("count", 1))

            # Sécurité si la clé et le result.id divergent
            if result_id != result_id_from_key:
                pass

            ingredients = extract_ingredients_from_recipe(recipe)

            normalized[result_id].append({
                "result_id": result_id,
                "result_count": result_count,
                "ingredients": ingredients,
                "raw": recipe
            })

    return dict(normalized)


def load_selected_recipe_choices(path: Path) -> Dict[str, int]:
    """
    Lit le fichier JSON qui indique, pour certains items, quel numero de craft garder.

    Format attendu :
    {
        "barrel": 12,
        "cake": 3
    }

    Le numero de craft est base sur l'ordre affiche dans le rapport
    `duplicate_recipe_results.txt`, donc il commence a 1.
    """
    if not path.exists():
        print(f"[Choix recettes] Fichier absent, aucun filtrage applique : {path}", flush=True)
        return {}

    data = load_json(path)
    if not isinstance(data, dict):
        raise ValueError(f"Le fichier {path} doit contenir un objet JSON nom -> numero de craft.")

    cleaned_choices = {}
    for item_name, recipe_number in data.items():
        # On verifie defensivement les types lus depuis le JSON.
        # Un JSON venant d'une edition manuelle peut contenir des erreurs.
        if not isinstance(item_name, str):
            continue
        try:
            cleaned_choices[item_name] = int(recipe_number)
        except (TypeError, ValueError):
            print(f"[Choix recettes] Valeur ignoree pour {item_name!r}: {recipe_number!r}", flush=True)

    print(f"[Choix recettes] {len(cleaned_choices)} selections chargees depuis {path}", flush=True)
    return cleaned_choices


def apply_selected_recipe_choices(
    recipes_by_result: Dict[int, List[Dict[str, Any]]],
    items_by_id: Dict[int, dict],
    selected_recipe_choices: Dict[str, int],
) -> Dict[int, List[Dict[str, Any]]]:
    """
    Filtre les recettes multiples pour ne garder que le craft choisi par l'utilisateur.

    Idee :
    - si un item a plusieurs recettes et qu'un numero est fourni dans
      `selected_recipe_choices`, on ne garde que cette recette
    - sinon, on laisse toutes les recettes disponibles

    Pourquoi le filtrage est fait ici ?
    - parce qu'on veut simplifier le moteur de calcul principal
    - le solveur continue ainsi de travailler sur "la liste des recettes
      autorisees", sans savoir si cette liste a ete reduite ou non
    """
    filtered_recipes = {}

    for result_item_id, recipe_list in recipes_by_result.items():
        # On retrouve le nom de l'item resultat pour pouvoir faire la
        # correspondance avec le fichier JSON des choix utilisateur.
        item = items_by_id.get(result_item_id)
        item_name = item["name"] if item else None

        # Si aucun choix n'existe pour cet item, ou s'il n'a qu'une seule
        # recette, on ne touche a rien.
        if not item_name or item_name not in selected_recipe_choices or len(recipe_list) <= 1:
            filtered_recipes[result_item_id] = recipe_list
            continue

        selected_recipe_number = selected_recipe_choices[item_name]
        selected_index = selected_recipe_number - 1

        # Les numerots affiches a l'utilisateur commencent a 1.
        # Les listes Python, elles, commencent a 0.
        if 0 <= selected_index < len(recipe_list):
            filtered_recipes[result_item_id] = [recipe_list[selected_index]]
            print(
                f"[Choix recettes] {item_name}: craft {selected_recipe_number} conserve sur {len(recipe_list)}",
                flush=True,
            )
        else:
            filtered_recipes[result_item_id] = recipe_list
            print(
                f"[Choix recettes] {item_name}: craft {selected_recipe_number} invalide, toutes les recettes gardees",
                flush=True,
            )

    return filtered_recipes


# =========================
# Outils de calcul
# =========================

class ResolutionError(Exception):
    """Exception personnalisee reservee aux erreurs de resolution."""
    pass


def counter_to_named_dict(counter: Counter, items_by_id: Dict[int, dict]) -> Dict[str, int]:
    """Transforme un Counter base sur des ids en dictionnaire base sur des noms."""
    result = {}
    for item_id, qty in sorted(counter.items(), key=lambda kv: items_by_id.get(kv[0], {}).get("name", str(kv[0]))):
        item = items_by_id.get(item_id)
        name = item["name"] if item else f"unknown_{item_id}"
        result[name] = int(qty)
    return result


def multiply_counter(counter: Counter, factor: int) -> Counter:
    """Multiplie toutes les quantites d'un Counter par un facteur."""
    return Counter({k: v * factor for k, v in counter.items()})


def merge_counters(a: Counter, b: Counter) -> Counter:
    """Fusionne deux compteurs en additionnant les quantites cle par cle."""
    result = Counter(a)
    result.update(b)
    return result


def total_base_cost(counter: Counter) -> int:
    """
    Mesure simple du "coût" :
    somme des quantités de ressources de base.
    """
    return sum(counter.values())


def convert_qty_with_ratio(qty: int, multiplier_num: int, multiplier_den: int) -> int:
    """Convertit une quantite selon un ratio, en arrondissant toujours au dessus."""
    return math.ceil(qty * multiplier_num / multiplier_den)


# =========================
# Résolution récursive
# =========================

class CraftAnalyzer:
    """
    Objet principal charge de resoudre les couts de craft.

    Pourquoi une classe ici ?
    - elle regroupe les donnees partagees par toutes les fonctions de calcul
    - elle evite de passer sans arret les memes variables en argument
    - elle permet de garder une memoisation interne propre
    """
    def __init__(
        self,
        items_by_id: Dict[int, dict],
        items_by_name: Dict[str, dict],
        recipes_by_result: Dict[int, List[Dict[str, Any]]],
        base_farmables_by_name: Set[str],
        allowed_items_by_name: Set[str],
    ):
        """
        Prepare toutes les structures utiles au calcul.

        On transforme ici les noms en ids tres tot, car les ids sont plus
        compacts et plus stables pour le calcul interne.
        """
        self.items_by_id = items_by_id
        self.items_by_name = items_by_name
        self.recipes_by_result = recipes_by_result

        # On convertit les noms "base farmables" en ids une fois pour toutes.
        self.base_farmable_ids = {
            self.items_by_name[name]["id"]
            for name in base_farmables_by_name
            if name in self.items_by_name
        }
        # Meme principe pour les ressources prioritaires.
        self.priority_base_ids = {
            self.items_by_name[name]["id"]
            for name in PRIORITY_BASE_ITEMS_BY_NAME
            if name in self.items_by_name
        }
        # Les items autorises viennent de la liste blanche du musee.
        self.allowed_item_ids = {
            self.items_by_name[name]["id"]
            for name in allowed_items_by_name
            if name in self.items_by_name
        }
        # Tout item non autorise est considere comme exclu.
        self.excluded_item_ids = set(self.items_by_id.keys()) - self.allowed_item_ids

        # Ce dictionnaire sert a ramener certaines formes vers une ressource
        # canonique, par exemple iron_block -> iron_ingot.
        self.reversible_alias_rules_by_id = {}
        for alias_name, (canonical_name, multiplier_num, multiplier_den) in REVERSIBLE_CANONICAL_RULES.items():
            alias_item = self.items_by_name.get(alias_name)
            canonical_item = self.items_by_name.get(canonical_name)
            if alias_item and canonical_item:
                self.reversible_alias_rules_by_id[alias_item["id"]] = (
                    canonical_item["id"],
                    multiplier_num,
                    multiplier_den,
                )

        # Mémoisation : (item_id, qty) -> résultat
        self.memo: Dict[Tuple[int, int], Dict[str, Any]] = {}

    def item_name(self, item_id: int) -> str:
        """Retourne le nom d'un item a partir de son id."""
        item = self.items_by_id.get(item_id)
        return item["name"] if item else f"unknown_{item_id}"

    def is_excluded(self, item_id: int) -> bool:
        """Indique si un item ne doit pas etre traite dans l'analyse."""
        return item_id in self.excluded_item_ids

    def is_base_farmable(self, item_id: int) -> bool:
        """Indique si l'on s'arrete sur cet item comme ressource de base."""
        return item_id in self.base_farmable_ids

    def has_recipe(self, item_id: int) -> bool:
        """Indique si l'item possede au moins une recette connue."""
        return item_id in self.recipes_by_result and len(self.recipes_by_result[item_id]) > 0

    def normalize_leaf_item(self, item_id: int, qty: int) -> Tuple[int, int]:
        # Normalise une ressource "feuille" vers sa forme canonique de recolte.
        # Exemple :
        # - iron_block -> iron_ingot
        # - iron_nugget -> iron_ingot
        # - stick -> oak_log
        rule = self.reversible_alias_rules_by_id.get(item_id)
        if not rule:
            return item_id, qty

        canonical_id, multiplier_num, multiplier_den = rule
        return canonical_id, convert_qty_with_ratio(qty, multiplier_num, multiplier_den)

    def normalize_counter(self, counter: Counter) -> Counter:
        """Applique la normalisation canonique a toutes les entrees d'un Counter."""
        normalized = Counter()
        for item_id, qty in counter.items():
            normalized_item_id, normalized_qty = self.normalize_leaf_item(item_id, qty)
            normalized[normalized_item_id] += normalized_qty
        return normalized

    def reversible_penalty_for_ingredients(self, ingredients: Counter) -> int:
        """Calcule une penalite pour les recettes qui utilisent des formes reversibles."""
        penalty = 0
        for ingredient_id, ingredient_qty in ingredients.items():
            if ingredient_id in self.reversible_alias_rules_by_id:
                penalty += ingredient_qty
        return penalty

    def priority_base_bonus(self, counter: Counter) -> int:
        """Mesure combien une solution s'exprime deja en ressources prioritaires."""
        return sum(qty for item_id, qty in counter.items() if item_id in self.priority_base_ids)

    def resolve_item(
        self,
        item_id: int,
        required_qty: int,
        visiting: Optional[Set[int]] = None
    ) -> Dict[str, Any]:
        """
        Fonction recursive principale du programme.

        Elle repond a la question :
        "Que faut-il au final pour obtenir `required_qty` exemplaires de cet item ?"

        La fonction retourne un dictionnaire avec :
        - `base_resources` : ce qu'il faut vraiment farmer
        - `unresolved` : ce qu'on n'a pas su decomposer
        - `excluded` : ce qui a ete ignore car hors liste
        - `recipe_used` : la recette choisie si l'item est craftable

        Le parametre `visiting` sert a detecter les cycles.
        Exemple de probleme sans cette protection :
        A demande B, B demande C, C redemande A.
        """
        if visiting is None:
            visiting = set()

        # La memoisation est un cache :
        # si on a deja calcule cet item pour cette quantite exacte,
        # on reutilise le resultat au lieu de recalculer tout l'arbre.
        memo_key = (item_id, required_qty)
        if memo_key in self.memo:
            return deepcopy(self.memo[memo_key])

        # Cas simple : demander 0 item ne coute rien.
        if required_qty <= 0:
            result = {
                "base_resources": Counter(),
                "unresolved": Counter(),
                "excluded": Counter(),
                "recipe_used": None
            }
            self.memo[memo_key] = deepcopy(result)
            return result

        # Exclu
        # Si l'item est exclu, on ne poursuit pas sa decomposition.
        if self.is_excluded(item_id):
            normalized_item_id, normalized_qty = self.normalize_leaf_item(item_id, required_qty)
            result = {
                "base_resources": Counter(),
                "unresolved": Counter(),
                "excluded": Counter({normalized_item_id: normalized_qty}),
                "recipe_used": None
            }
            self.memo[memo_key] = deepcopy(result)
            return result

        # Ressource de base
        # Si l'item est une ressource de base, on considere que c'est
        # la fin de la chaine de craft.
        if self.is_base_farmable(item_id):
            normalized_item_id, normalized_qty = self.normalize_leaf_item(item_id, required_qty)
            result = {
                "base_resources": Counter({normalized_item_id: normalized_qty}),
                "unresolved": Counter(),
                "excluded": Counter(),
                "recipe_used": None
            }
            self.memo[memo_key] = deepcopy(result)
            return result

        # Anti-cycle
        # Protection anti-boucle :
        # si on revoit le meme item dans la pile courante de recursion,
        # on coupe pour eviter une boucle infinie.
        if item_id in visiting:
            normalized_item_id, normalized_qty = self.normalize_leaf_item(item_id, required_qty)
            result = {
                "base_resources": Counter(),
                "unresolved": Counter({normalized_item_id: normalized_qty}),
                "excluded": Counter(),
                "recipe_used": None
            }
            self.memo[memo_key] = deepcopy(result)
            return result

        # Pas de recette => non résolu => on considère qu'on doit le farmer tel quel
        # ou au moins le signaler.
        # Si aucune recette n'existe, on ne peut pas aller plus loin.
        if not self.has_recipe(item_id):
            normalized_item_id, normalized_qty = self.normalize_leaf_item(item_id, required_qty)
            result = {
                "base_resources": Counter(),
                "unresolved": Counter({normalized_item_id: normalized_qty}),
                "excluded": Counter(),
                "recipe_used": None
            }
            self.memo[memo_key] = deepcopy(result)
            return result

        # On copie l'ensemble pour ne pas modifier l'objet recu depuis l'appel
        # parent. C'est une bonne pratique quand on veut garder un etat local
        # propre a cette branche de recursion.
        visiting = set(visiting)
        visiting.add(item_id)

        candidate_results = []

        # Un item peut avoir plusieurs recettes.
        # On va toutes les tester, puis choisir la meilleure a la fin.
        for recipe in self.recipes_by_result[item_id]:
            result_count = recipe["result_count"]
            ingredients = recipe["ingredients"]

            if result_count <= 0:
                continue

            # Nombre de crafts necessaires pour atteindre la quantite demandee.
            crafts_needed = math.ceil(required_qty / result_count)

            aggregate_base = Counter()
            aggregate_unresolved = Counter()
            aggregate_excluded = Counter()

            # On resout ensuite chaque ingredient de la recette.
            for ingredient_id, ingredient_qty in ingredients.items():
                needed_ingredient_qty = ingredient_qty * crafts_needed
                sub = self.resolve_item(ingredient_id, needed_ingredient_qty, visiting=visiting)

                aggregate_base.update(sub["base_resources"])
                aggregate_unresolved.update(sub["unresolved"])
                aggregate_excluded.update(sub["excluded"])

            # Une fois toute la recette resolue, on normalise les formes
            # reversibles pour exprimer le besoin dans une forme canonique.
            aggregate_base = self.normalize_counter(aggregate_base)
            aggregate_unresolved = self.normalize_counter(aggregate_unresolved)
            aggregate_excluded = self.normalize_counter(aggregate_excluded)
            reversible_penalty = self.reversible_penalty_for_ingredients(ingredients)

            # On stocke la solution candidate complete pour comparaison.
            candidate_results.append({
                "base_resources": aggregate_base,
                "unresolved": aggregate_unresolved,
                "excluded": aggregate_excluded,
                "recipe_used": {
                    "result_id": recipe["result_id"],
                    "result_count": recipe["result_count"],
                    "ingredients": dict(recipe["ingredients"]),
                    "crafts_needed": crafts_needed
                },
                "reversible_penalty": reversible_penalty,
                "score": (
                    sum(aggregate_unresolved.values()),  # d'abord minimiser les non résolus
                    total_base_cost(aggregate_base),      # puis minimiser les ressources de base
                    sum(aggregate_excluded.values())      # puis les exclus
                )
            })

        if not candidate_results:
            normalized_item_id, normalized_qty = self.normalize_leaf_item(item_id, required_qty)
            result = {
                "base_resources": Counter(),
                "unresolved": Counter({normalized_item_id: normalized_qty}),
                "excluded": Counter(),
                "recipe_used": None
            }
            self.memo[memo_key] = deepcopy(result)
            return result

        # On choisit ensuite la "meilleure" recette selon plusieurs criteres
        # classes par ordre d'importance.
        best = min(
            candidate_results,
            key=lambda x: (
                sum(x["unresolved"].values()),
                x["reversible_penalty"],
                -self.priority_base_bonus(x["base_resources"]),
                total_base_cost(x["base_resources"]),
                sum(x["excluded"].values()),
            ),
        )

        result = {
            "base_resources": best["base_resources"],
            "unresolved": best["unresolved"],
            "excluded": best["excluded"],
            "recipe_used": best["recipe_used"]
        }
        self.memo[memo_key] = deepcopy(result)
        return result

    def target_quantity_for_item(self, item_id: int) -> int:
        """
        Calcule la quantite cible pour un item.

        Exemple :
        - stackSize = 64
        - 27 slots dans un coffre
        => cible = 1728
        """
        item = self.items_by_id[item_id]
        stack_size = int(item.get("stackSize", 64))
        if stack_size <= 0:
            stack_size = 1
        return CHEST_SLOTS * stack_size


# =========================
# Analyse globale
# =========================

def analyze_all_items():
    """
    Fonction "chef d'orchestre" du programme.

    C'est elle qui enchaine toutes les etapes dans le bon ordre :
    1. verifier la presence des fichiers
    2. charger les donnees
    3. preparer l'analyseur
    4. parcourir chaque item autorise
    5. cumuler les resultats globaux
    6. ecrire les fichiers de sortie
    """
    if not ITEMS_FILE.exists():
        raise FileNotFoundError(f"Fichier introuvable: {ITEMS_FILE}")
    if not RECIPES_FILE.exists():
        raise FileNotFoundError(f"Fichier introuvable: {RECIPES_FILE}")
    if not MUSEUM_EXCEL_FILE.exists():
        raise FileNotFoundError(f"Fichier introuvable: {MUSEUM_EXCEL_FILE}")

    # Lecture des deux gros fichiers JSON d'entree.
    print(f"[Init] Chargement de {ITEMS_FILE}...", flush=True)
    items_data = load_json(ITEMS_FILE)
    print(f"[Init] {len(items_data)} items charges.", flush=True)

    print(f"[Init] Chargement de {RECIPES_FILE}...", flush=True)
    recipes_data = load_json(RECIPES_FILE)
    print(f"[Init] {len(recipes_data)} entrees de recettes chargees.", flush=True)

    # Ici on prepare plusieurs "vues" des memes donnees :
    # - par id
    # - par nom
    # - par recettes normalisees
    items_by_id, items_by_name = build_item_maps(items_data)
    item_lookup_map = build_item_lookup_map(items_data)
    recipes_by_result = normalize_recipes(recipes_data)

    # On charge ici un fichier de preferences utilisateur indiquant, pour
    # certains items a recettes multiples, quel craft doit etre conserve.
    # Cela permet d'eviter les doublons de variantes et d'imposer des choix
    # plus coherents avec le projet (exemple : une essence de bois precise).
    selected_recipe_choices = load_selected_recipe_choices(SELECTED_RECIPES_FILE)

    # Une fois les recettes normalisees, on applique le filtre.
    # A partir de ce point, le moteur de calcul travaille uniquement avec
    # les recettes autorisees apres selection.
    recipes_by_result = apply_selected_recipe_choices(
        recipes_by_result=recipes_by_result,
        items_by_id=items_by_id,
        selected_recipe_choices=selected_recipe_choices,
    )

    # Ces deux ensembles servent a enrichir l'export Excel avec deux questions :
    # - l'item apparait-il comme resultat d'une recette ?
    # - l'item apparait-il comme ingredient d'une recette ?
    recipe_result_items_by_name = set()
    recipe_ingredient_items_by_name = set()
    for result_item_id, recipe_list in recipes_by_result.items():
        result_item = items_by_id.get(result_item_id)
        if result_item:
            recipe_result_items_by_name.add(result_item["name"])

        for recipe in recipe_list:
            for ingredient_item_id in recipe["ingredients"].keys():
                ingredient_item = items_by_id.get(int(ingredient_item_id))
                if ingredient_item:
                    recipe_ingredient_items_by_name.add(ingredient_item["name"])

    allowed_items_by_name = load_allowed_items_from_excel(
        excel_path=MUSEUM_EXCEL_FILE,
        items_by_name=items_by_name,
        item_lookup_map=item_lookup_map,
    )

    # L'objet `analyzer` contient toutes les regles et toutes les donnees
    # partagees par la resolution recursive.
    analyzer = CraftAnalyzer(
        items_by_id=items_by_id,
        items_by_name=items_by_name,
        recipes_by_result=recipes_by_result,
        base_farmables_by_name=BASE_FARMABLES_BY_NAME,
        allowed_items_by_name=allowed_items_by_name,
    )

    grand_total_base = Counter()
    grand_total_unresolved = Counter()
    grand_total_excluded = Counter()

    per_item_details = {}

    # On trie les items pour avoir toujours le meme ordre de sortie.
    # C'est tres utile pour comparer deux executions ou deux commits.
    all_item_ids = sorted(items_by_id.keys(), key=lambda i: items_by_id[i]["name"])
    item_ids_to_process = [item_id for item_id in all_item_ids if item_id not in analyzer.excluded_item_ids]
    total_items_to_process = len(item_ids_to_process)
    print(f"[Init] {total_items_to_process} items a analyser.", flush=True)

    # Les items absents de la liste musee ne sont pas traites, donc on les
    # ajoute explicitement au recapitulatif final pour qu'ils apparaissent bien.
    excluded_items_summary = Counter(
        {
            item_id: analyzer.target_quantity_for_item(item_id)
            for item_id in sorted(analyzer.excluded_item_ids, key=lambda i: items_by_id[i]["name"])
        }
    )

    # Ces ajustements sont volontairement appliques a la fin, car ils
    # representent des besoins "hors calcul automatique".
    manual_total_adjustments = Counter()
    for item_name, qty_to_add in MANUAL_TOTAL_ADJUSTMENTS_BY_NAME.items():
        item = items_by_name.get(item_name)
        if not item:
            print(f"[Ajustement] Item introuvable ignore: {item_name}", flush=True)
            continue
        manual_total_adjustments[item["id"]] += qty_to_add
        print(f"[Ajustement] {item_name} +{qty_to_add}", flush=True)

    # Boucle principale :
    # on calcule independamment le cout de chaque item autorise.
    for index, item_id in enumerate(item_ids_to_process, start=1):
        item = items_by_id[item_id]
        item_name = item["name"]

        target_qty = analyzer.target_quantity_for_item(item_id)
        print(
            f"[Analyse] Item {index}/{total_items_to_process} - ligne logique {index}: {item_name} (quantite cible: {target_qty})",
            flush=True,
        )
        resolution = analyzer.resolve_item(item_id, target_qty)

        # Chaque resolution renvoie plusieurs compteurs que l'on additionne
        # ensuite dans les totaux globaux.
        grand_total_base.update(resolution["base_resources"])
        grand_total_unresolved.update(resolution["unresolved"])
        grand_total_excluded.update(resolution["excluded"])

        # On stocke aussi un detail complet pour cet item precis.
        per_item_details[item_name] = {
            "item_id": item_id,
            "display_name": item.get("displayName", item_name),
            "stack_size": item.get("stackSize", 64),
            "target_quantity": target_qty,
            "base_resources": counter_to_named_dict(resolution["base_resources"], items_by_id),
            "unresolved": counter_to_named_dict(resolution["unresolved"], items_by_id),
            "excluded": counter_to_named_dict(resolution["excluded"], items_by_id),
            "recipe_used": format_recipe_used(resolution["recipe_used"], items_by_id),
        }

        print_progress(index, total_items_to_process)

    # On prepare une structure JSON propre, facile a relire ou a reutiliser
    # depuis un autre script plus tard.
    summary = {
        "config": {
            "chest_slots": CHEST_SLOTS,
            "base_farmables": sorted(BASE_FARMABLES_BY_NAME),
            "allowed_items": sorted(allowed_items_by_name),
            "manual_total_adjustments": dict(MANUAL_TOTAL_ADJUSTMENTS_BY_NAME),
            "selected_recipe_choices": selected_recipe_choices,
        },
        "global_totals": {
            "base_resources": counter_to_named_dict(
                merge_counters(grand_total_base, manual_total_adjustments),
                items_by_id,
            ),
            "unresolved": counter_to_named_dict(grand_total_unresolved, items_by_id),
            "excluded": counter_to_named_dict(
                merge_counters(grand_total_excluded, excluded_items_summary),
                items_by_id,
            ),
        },
        "per_item_details": per_item_details,
    }

    # Ecriture du JSON detaille.
    with OUTPUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)

    # Cette exportation Excel est faite a la fin de CHAQUE execution.
    # On depend du resume final `summary`, donc il faut attendre que tous les
    # calculs soient termines avant de remplir les colonnes H et I.
    export_museum_excel_with_totals(
        source_excel_path=MUSEUM_EXCEL_FILE,
        output_excel_path=MUSEUM_OUTPUT_EXCEL_FILE,
        items_data=items_data,
        summary=summary,
        recipe_result_items_by_name=recipe_result_items_by_name,
        recipe_ingredient_items_by_name=recipe_ingredient_items_by_name,
    )

    # Ecriture des autres sorties plus "humaines".
    write_recipe_tree_summary(
        analyzer=analyzer,
        item_ids_to_process=item_ids_to_process,
    )

    write_human_readable_summary(
        items_by_id=items_by_id,
        grand_total_base=merge_counters(grand_total_base, manual_total_adjustments),
        grand_total_unresolved=grand_total_unresolved,
        grand_total_excluded=merge_counters(grand_total_excluded, excluded_items_summary),
        per_item_details=per_item_details,
    )

    print(f"Analyse terminée.")
    print(f"- Résumé JSON : {OUTPUT_JSON}")
    print(f"- Résumé texte : {OUTPUT_TXT}")


def format_recipe_used(recipe_used: Optional[Dict[str, Any]], items_by_id: Dict[int, dict]) -> Optional[Dict[str, Any]]:
    """
    Transforme la recette choisie en version lisible avec des noms d'items.

    En interne, les recettes utilisent surtout des ids.
    Pour l'affichage, les noms sont beaucoup plus confortables.
    """
    if recipe_used is None:
        return None

    ingredients_named = {}
    for item_id, qty in recipe_used["ingredients"].items():
        name = items_by_id.get(int(item_id), {}).get("name", f"unknown_{item_id}")
        ingredients_named[name] = qty

    return {
        "result_id": recipe_used["result_id"],
        "result_name": items_by_id.get(recipe_used["result_id"], {}).get("name", f"unknown_{recipe_used['result_id']}"),
        "result_count": recipe_used["result_count"],
        "crafts_needed": recipe_used["crafts_needed"],
        "ingredients": ingredients_named,
    }


def export_museum_excel_with_totals(
    source_excel_path: Path,
    output_excel_path: Path,
    items_data: List[Dict[str, Any]],
    summary: Dict[str, Any],
    recipe_result_items_by_name: Set[str],
    recipe_ingredient_items_by_name: Set[str],
) -> None:
    """
    Cree une copie du fichier Excel du musee et y ajoute deux colonnes calculees.

    Colonnes ajoutees :
    - H : "Total à farm"
      total global a farmer pour cet item
      (on additionne ici `base_resources` et `unresolved`, car certains
      items minables ou non craftables peuvent etre calcules mais classes
      dans `unresolved` par le moteur)
    - I : "Total cible item"
      quantite cible de cet item lui-meme
    - J : "OnRecipes_Res"
      `true` si l'item apparait comme resultat dans `recipes.json`
    - K : "OnRecipes_Ingredient"
      `true` si l'item apparait comme ingredient dans `recipes.json`

    Pourquoi faire cette exportation a la fin ?
    - parce que ces valeurs n'existent qu'une fois les calculs termines
    - on a besoin du `summary` final pour remplir les cellules

    Pourquoi faire une copie plutot que modifier l'original ?
    - pour conserver le fichier source intact
    - pour separer clairement les donnees d'entree et les donnees generees
    """
    # On reconstruit une table de correspondance texte -> item_name pour faire
    # le lien entre la colonne `english_name` du musee et les noms internes.
    item_lookup_map = build_item_lookup_map(items_data)
    global_base_resources = summary.get("global_totals", {}).get("base_resources", {})
    global_unresolved_resources = summary.get("global_totals", {}).get("unresolved", {})
    per_item_details = summary.get("per_item_details", {})

    print(f"[Excel Export] Creation de {output_excel_path}...", flush=True)

    # Ici on ouvre le classeur en mode normal (pas read_only), car on veut
    # modifier des cellules puis sauvegarder un nouveau fichier.
    workbook = load_workbook(source_excel_path)
    try:
        sheet_name = "Sheet1" if "Sheet1" in workbook.sheetnames else workbook.sheetnames[0]
        worksheet = workbook[sheet_name]

        # On ecrit les en-tetes des nouvelles colonnes.
        worksheet.cell(row=1, column=8).value = "Total Pour Items Craftables"
        worksheet.cell(row=1, column=9).value = "Total cible item"
        worksheet.cell(row=1, column=10).value = "OnRecipes_Res"
        worksheet.cell(row=1, column=11).value = "OnRecipes_Ingredient"

        # On parcourt toutes les lignes de donnees du fichier musee.
        for row_number in range(2, worksheet.max_row + 1):
            english_name = worksheet.cell(row=row_number, column=7).value

            total_to_farm_cell = worksheet.cell(row=row_number, column=8)
            target_total_cell = worksheet.cell(row=row_number, column=9)
            on_recipes_result_cell = worksheet.cell(row=row_number, column=10)
            on_recipes_ingredient_cell = worksheet.cell(row=row_number, column=11)

            # Si la ligne n'a pas de nom anglais, on met 0 dans les colonnes.
            if not english_name:
                total_to_farm_cell.value = 0
                target_total_cell.value = 0
                on_recipes_result_cell.value = "false"
                on_recipes_ingredient_cell.value = "false"
                continue

            raw_excel_item_name = str(english_name).strip()
            normalized_excel_item_name = normalize_item_lookup_key(raw_excel_item_name)
            resolved_item_name = (
                item_lookup_map.get(raw_excel_item_name)
                or item_lookup_map.get(normalized_excel_item_name)
            )

            # Si aucun item n'est trouve, on met 0.
            if not resolved_item_name:
                total_to_farm_cell.value = 0
                target_total_cell.value = 0
                on_recipes_result_cell.value = "false"
                on_recipes_ingredient_cell.value = "false"
                continue

            # La colonne H utilise le total global calcule pour cet item.
            # On additionne :
            # - `base_resources` : vraies ressources de base reconnues
            # - `unresolved` : besoins calcules mais non classes comme base
            # Cela corrige par exemple des cas comme `cobbled_deepslate`,
            # qui est bien comptabilise par le calcul mais pouvait apparaitre
            # dans `unresolved` plutot que dans `base_resources`.
            total_to_farm_cell.value = int(
                global_base_resources.get(resolved_item_name, 0)
                + global_unresolved_resources.get(resolved_item_name, 0)
            )

            # La colonne I utilise la quantite cible de l'item lui-meme.
            item_detail = per_item_details.get(resolved_item_name, {})
            target_total_cell.value = int(item_detail.get("target_quantity", 0))

            # Colonnes J et K :
            # - resultat de recette ?
            # - ingredient de recette ?
            on_recipes_result_cell.value = "true" if resolved_item_name in recipe_result_items_by_name else "false"
            on_recipes_ingredient_cell.value = "true" if resolved_item_name in recipe_ingredient_items_by_name else "false"

        workbook.save(output_excel_path)
    finally:
        workbook.close()


def build_recipe_tree_lines(
    analyzer: CraftAnalyzer,
    item_id: int,
    required_qty: int = 1,
    depth: int = 0,
    visiting: Optional[Set[int]] = None,
) -> List[str]:
    if visiting is None:
        visiting = set()

    item_name = analyzer.item_name(item_id)
    if depth == 0:
        lines = [item_name]
    else:
        lines = [f"{'    ' * depth}└── {item_name} x{required_qty}"]

    if item_id in visiting:
        lines[-1] += " [cycle]"
        return lines

    if analyzer.is_excluded(item_id):
        lines[-1] += " [exclu]"
        return lines

    if analyzer.is_base_farmable(item_id):
        return lines

    if not analyzer.has_recipe(item_id):
        lines[-1] += " [non resolu]"
        return lines

    resolution = analyzer.resolve_item(item_id, required_qty)
    recipe_used = resolution.get("recipe_used")
    if not recipe_used:
        return lines

    next_visiting = set(visiting)
    next_visiting.add(item_id)

    for ingredient_id, ingredient_qty in sorted(
        recipe_used["ingredients"].items(),
        key=lambda kv: analyzer.item_name(int(kv[0])),
    ):
        child_required_qty = ingredient_qty * recipe_used["crafts_needed"]
        lines.extend(
            build_recipe_tree_lines(
                analyzer=analyzer,
                item_id=int(ingredient_id),
                required_qty=child_required_qty,
                depth=depth + 1,
                visiting=next_visiting,
            )
        )

    return lines


def build_recipe_tree_lines_v2(
    analyzer: CraftAnalyzer,
    item_id: int,
    required_qty: int = 1,
    depth: int = 0,
    visiting: Optional[Set[int]] = None,
) -> List[str]:
    """
    Construit recursivement un petit arbre texte pour afficher une recette.

    Cette fonction est surtout orientee "lecture humaine" et non performance.
    Elle sert a visualiser les dependances entre crafts.
    """
    if visiting is None:
        visiting = set()

    item_name = analyzer.item_name(item_id)
    produced_qty = required_qty
    recipe_used = None

    if analyzer.has_recipe(item_id):
        resolution = analyzer.resolve_item(item_id, required_qty)
        recipe_used = resolution.get("recipe_used")
        if recipe_used:
            produced_qty = recipe_used["result_count"] * recipe_used["crafts_needed"]

    if depth == 0:
        lines = [f"{item_name} x{produced_qty} (pour {required_qty} craft)"]
    else:
        lines = [f"{'    ' * depth}\\-- {item_name} x{required_qty}"]

    if item_id in visiting:
        lines[-1] += " [cycle]"
        return lines

    if analyzer.is_excluded(item_id):
        lines[-1] += " [exclu]"
        return lines

    if analyzer.is_base_farmable(item_id):
        return lines

    if not analyzer.has_recipe(item_id):
        lines[-1] += " [non resolu]"
        return lines

    if not recipe_used:
        return lines

    next_visiting = set(visiting)
    next_visiting.add(item_id)

    for ingredient_id, ingredient_qty in sorted(
        recipe_used["ingredients"].items(),
        key=lambda kv: analyzer.item_name(int(kv[0])),
    ):
        child_required_qty = ingredient_qty * recipe_used["crafts_needed"]
        lines.extend(
            build_recipe_tree_lines_v2(
                analyzer=analyzer,
                item_id=int(ingredient_id),
                required_qty=child_required_qty,
                depth=depth + 1,
                visiting=next_visiting,
            )
        )

    return lines


def write_recipe_tree_summary(
    analyzer: CraftAnalyzer,
    item_ids_to_process: List[int],
):
    """Ecrit le fichier texte qui contient un arbre de craft par item."""
    lines = [
        "=== ARBORESCENCE DES RECETTES ===",
        "Chaque arbre correspond a 1 craft de l'item.",
    ]

    for item_id in item_ids_to_process:
        lines.append("")
        lines.extend(build_recipe_tree_lines_v2(analyzer=analyzer, item_id=item_id, required_qty=1))

    with OUTPUT_TREE_TXT.open("w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_human_readable_summary(
    items_by_id: Dict[int, dict],
    grand_total_base: Counter,
    grand_total_unresolved: Counter,
    grand_total_excluded: Counter,
    per_item_details: Dict[str, Any],
):
    """
    Ecrit un grand fichier texte lisible pour un humain.

    Le but n'est pas d'etre compact, mais d'etre facile a parcourir :
    - d'abord les totaux globaux
    - ensuite les cas non resolus / exclus
    - enfin le detail item par item
    """
    lines = []

    lines.append("=== TOTAL GLOBAL DES RESSOURCES À FARMER ===")
    if grand_total_base:
        for item_id, qty in sorted(grand_total_base.items(), key=lambda kv: items_by_id[kv[0]]["name"]):
            lines.append(f"- {items_by_id[item_id]['name']}: {qty}")
    else:
        lines.append("(aucune)")

    lines.append("")
    lines.append("=== ITEMS NON RÉSOLUS ===")
    if grand_total_unresolved:
        for item_id, qty in sorted(grand_total_unresolved.items(), key=lambda kv: items_by_id.get(kv[0], {}).get("name", str(kv[0]))):
            name = items_by_id.get(item_id, {}).get("name", f"unknown_{item_id}")
            lines.append(f"- {name}: {qty}")
    else:
        lines.append("(aucun)")

    lines.append("")
    lines.append("=== ITEMS EXCLUS ===")
    if grand_total_excluded:
        for item_id, qty in sorted(grand_total_excluded.items(), key=lambda kv: items_by_id.get(kv[0], {}).get("name", str(kv[0]))):
            name = items_by_id.get(item_id, {}).get("name", f"unknown_{item_id}")
            lines.append(f"- {name}: {qty}")
    else:
        lines.append("(aucun)")

    lines.append("")
    lines.append("=== DÉTAIL PAR ITEM ===")

    for item_name in sorted(per_item_details.keys()):
        detail = per_item_details[item_name]
        lines.append("")
        lines.append(f"[{item_name}]")
        lines.append(f"  - target_quantity: {detail['target_quantity']}")
        lines.append(f"  - stack_size: {detail['stack_size']}")

        if detail["recipe_used"]:
            lines.append("  - recipe_used:")
            lines.append(f"      result_count: {detail['recipe_used']['result_count']}")
            lines.append(f"      crafts_needed: {detail['recipe_used']['crafts_needed']}")
            lines.append("      ingredients:")
            for ing_name, ing_qty in sorted(detail["recipe_used"]["ingredients"].items()):
                lines.append(f"        - {ing_name}: {ing_qty}")
        else:
            lines.append("  - recipe_used: none")

        lines.append("  - base_resources:")
        if detail["base_resources"]:
            for res_name, res_qty in sorted(detail["base_resources"].items()):
                lines.append(f"      - {res_name}: {res_qty}")
        else:
            lines.append("      (none)")

        lines.append("  - unresolved:")
        if detail["unresolved"]:
            for res_name, res_qty in sorted(detail["unresolved"].items()):
                lines.append(f"      - {res_name}: {res_qty}")
        else:
            lines.append("      (none)")

        lines.append("  - excluded:")
        if detail["excluded"]:
            for res_name, res_qty in sorted(detail["excluded"].items()):
                lines.append(f"      - {res_name}: {res_qty}")
        else:
            lines.append("      (none)")

    with OUTPUT_TXT.open("w", encoding="utf-8") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    analyze_all_items()
